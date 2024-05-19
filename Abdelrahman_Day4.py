from openpyxl import load_workbook
import socket

'''
workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "hello"
sheet["B1"] = "world!"

workbook.save(filename="hello.xlsx")
'''
'''
workbook= load_workbook('hello.xlsx')

worksheet = workbook.active

# Read data from specific cells
value_A1 = worksheet['A1'].value
value_B2 = worksheet.cell(row=1, column=2).value

print("Value in cell A1:", value_A1)
print("Value in cell B2:", value_B2)

'''
servers = [
    "192.168.1.1",
    "192.168.1.2",
    "192.168.1.3" ,
    "192.168.1.4"
]

def check_port_status(ip , port):
    try:
        
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        socket.setdefaulttimeout(1)        
        result = s.connect_ex((ip, port))


        if (result == 0):
            return "Port  is open"
        else:
            return "Port  is closed"
        
    finally:
        s.close()
for server in servers:
    print(f"Checking ports on server {server}:")
    for port in range(75, 85):
        status = check_port_status(server, port)
        print(f"Port {port}: {status}")
    print()